"""
file_engine/excel_writer.py
=============================
เฟส 2 ของระบบ (แยกจากเฟส 1 ที่รัน CropWat โดยสิ้นเชิง): สแกนโฟลเดอร์ .txt ที่มีอยู่
จริงทั้งหมด parse แล้วเขียนทับ sheet "Result" ใหม่ทั้งแผ่นใน Excel master — เรียกซ้ำ
ได้ทุกเมื่อโดยไม่ต้องรอเฟส 1 รันจนจบ (อ่านจากไฟล์ .txt ที่มีอยู่ ณ ตอนนั้นเท่านั้น
ไม่ได้ผูกกับ state การรันของเฟส 1 เลย)

ตอนนี้สร้างเฉพาะ sheet "Result" ก่อน (ยืนยันกับผู้ใช้แล้ว) ไม่แตะ sheet อื่นถ้ามีอยู่
แล้วในไฟล์ปลายทาง (Result-1 / หมายเหตุ ทำทีหลัง)

โครงสร้าง sheet ที่สร้าง (11 แถวต่อปี ยืนยันจากไฟล์ตัวอย่างจริง):
  - แถวหัวตาราง: คอลัมน์ A=ปี, B=ชื่อตัวแปร, C เป็นต้นไป=วันปลูกที่ทดลอง (union ของ
    ทุกวันปลูกที่เจอในไฟล์ .txt ทั้งหมด เรียงตามปฏิทิน mm/dd ใช้ชุดคอลัมน์เดียวกัน
    ทุกปี — ปีไหนไม่มีไฟล์สำหรับวันปลูกไหนก็เว้นว่างไว้)
  - ต่อปี 11 แถว: % yield reduction, Actual/Potential water use by crop,
    Total/Effective rainfall, Moist deficit at harvest, Actual irrigation
    requirement, Reductions in Etc Stage A-D
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from file_engine.txt_parser import ParsedCandidateResult, TxtParseError, parse_txt

logger = logging.getLogger("excel_writer")

METRIC_ROWS = [
    ("ค่า % yield reduction", lambda r: r.yield_reduction_pct),
    ("Actual water use by crop (mm)", lambda r: r.actual_water_use_mm),
    ("Potential water use by crop (mm)", lambda r: r.potential_water_use_mm),
    ("Total rainfall (mm)", lambda r: r.total_rainfall_mm),
    ("Effective rainfall (mm)", lambda r: r.effective_rainfall_mm),
    ("Moist deficit at harvest (mm)", lambda r: r.moist_deficit_at_harvest_mm),
    ("Actual irrigation requirement (mm)", lambda r: r.actual_irrigation_requirement_mm),
    ("Reductions in Etc (%) -Stage A", lambda r: r.reduction_etc_stage_a),
    ("Reductions in Etc (%) -Stage B", lambda r: r.reduction_etc_stage_b),
    ("Reductions in Etc (%) -Stage C", lambda r: r.reduction_etc_stage_c),
    ("Reductions in Etc (%) -Stage D", lambda r: r.reduction_etc_stage_d),
]

# ชื่อไฟล์ .txt ที่ automation engine สร้าง (ดู export_results ใน cropwat_engine.py):
# "{ปี}_{MMDD}.txt" เช่น "1981_0401.txt"
FILENAME_RE = re.compile(r"^(?P<year>\d{4})_(?P<mmdd>\d{4})\.txt$")


def _sort_key(mmdd: str) -> tuple[int, int]:
    return (int(mmdd[:2]), int(mmdd[2:]))


def _format_mmdd(mmdd: str) -> str:
    return f"{mmdd[2:]}/{mmdd[:2]}"


def _collect_results(txt_dir: Path) -> dict[int, dict[str, ParsedCandidateResult]]:
    """สแกน txt_dir หาไฟล์ .txt ทั้งหมด parse ทีละไฟล์ — ไฟล์ที่ parse ไม่ได้แค่
    ข้าม + log warning ไป ไม่ทำให้ทั้งเฟส 2 ล้ม (เหมือนหลักการเดียวกับเฟส 1)"""
    parsed: dict[int, dict[str, ParsedCandidateResult]] = {}
    for path in sorted(Path(txt_dir).glob("*.txt")):
        m = FILENAME_RE.match(path.name)
        if not m:
            logger.warning("ข้ามไฟล์ชื่อไม่ตรงรูปแบบที่คาด: %s", path.name)
            continue
        year = int(m.group("year"))
        mmdd = m.group("mmdd")
        try:
            result = parse_txt(path)
        except TxtParseError as exc:
            logger.warning("parse ไฟล์ %s ไม่สำเร็จ: %s", path.name, exc)
            continue
        parsed.setdefault(year, {})[mmdd] = result
    return parsed


def _load_or_create_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)  # ลบ sheet เริ่มต้น "Sheet" ที่ openpyxl สร้างมาให้เฉยๆ
    return wb


def build_result_sheet(txt_dir: Path, output_xlsx: Path) -> int:
    """สร้าง/เขียนทับ sheet 'Result' ใน output_xlsx จากไฟล์ .txt ทั้งหมดใน txt_dir
    คืนค่าจำนวนปีที่เขียนสำเร็จ"""
    parsed = _collect_results(txt_dir)
    if not parsed:
        raise FileNotFoundError(f"ไม่พบไฟล์ .txt ที่ parse ได้เลยในโฟลเดอร์: {txt_dir}")

    all_mmdd = sorted({mmdd for year_data in parsed.values() for mmdd in year_data}, key=_sort_key)
    years = sorted(parsed)

    wb = _load_or_create_workbook(Path(output_xlsx))
    if "Result" in wb.sheetnames:
        del wb["Result"]
    ws = wb.create_sheet("Result", 0)

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    bold = Font(bold=True)

    ws.cell(row=1, column=1, value="ปี").font = bold
    ws.cell(row=1, column=2, value="ตัวแปร").font = bold
    for col_idx, mmdd in enumerate(all_mmdd, start=3):
        cell = ws.cell(row=1, column=col_idx, value=_format_mmdd(mmdd))
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for year in years:
        year_data = parsed[year]
        first_row_of_block = row_idx
        for label, getter in METRIC_ROWS:
            ws.cell(row=row_idx, column=2, value=label)
            for col_idx, mmdd in enumerate(all_mmdd, start=3):
                result = year_data.get(mmdd)
                if result is not None:
                    ws.cell(row=row_idx, column=col_idx, value=getter(result))
            row_idx += 1
        ws.cell(row=first_row_of_block, column=1, value=year).font = bold
        ws.merge_cells(
            start_row=first_row_of_block, start_column=1, end_row=row_idx - 1, end_column=1
        )
        ws.cell(row=first_row_of_block, column=1).alignment = Alignment(vertical="center")

    ws.column_dimensions["B"].width = 32
    for col_idx in range(3, 3 + len(all_mmdd)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
    ws.freeze_panes = "C2"

    Path(output_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    return len(years)
